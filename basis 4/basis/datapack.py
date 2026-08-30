"""Generic datapack builder -- the agent-under-test, for ANY company.

Builds an IC-style Excel datapack from a company's EDGAR cache, following the
Anthropic investment-banking plugin's conventions: formulas for every
calculation (its RULE 6), blue inputs / black formulas / green cross-sheet
links, $M with units declared, years as text, accounting negatives, source
citations, exec summary with a linked snapshot.

Deliberately faithful to those conventions, including the ones that produce
auditable defects (e.g. a summary that links to a legitimately-blank growth
cell -- see FAILURES.md #9). The auditor's job is to catch what the
conventions let through; the builder's job is not to secretly dodge them.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name="Arial", size=11, color="0000FF")
BLACK = Font(name="Arial", size=11)
GREEN = Font(name="Arial", size=11, color="008000")
BOLD = Font(name="Arial", size=11, bold=True)
NOTE = Font(name="Arial", size=9, color="5B5B66")
MONEY = "$#,##0.0;($#,##0.0)"
PCT = "0.0%;(0.0%)"
DAYS = "#,##0.0"
RIGHT = Alignment(horizontal="right")
TOP = Border(top=Side(style="thin"))
DBL = Border(bottom=Side(style="double"))

N_YEARS = 5


def build_datapack(cache_path: str | Path, out_path: str | Path) -> Path:
    cache = json.loads(Path(cache_path).read_text())
    F, meta = cache["facts"], cache["meta"]
    R = meta["resolved"]
    periods = meta["fiscal_periods"]
    fys = sorted(periods)[-N_YEARS:]
    n = len(fys)
    cols = [get_column_letter(2 + i) for i in range(n)]
    notes_col = get_column_letter(2 + n)
    ticker = meta.get("ticker", "?")
    src = f"SEC EDGAR XBRL, 10-K filings (retrieved {meta.get('retrieved','?')})"

    def has(logical):
        return logical in R and all(periods[fy] in F[R[logical]] for fy in fys)

    def vals(logical):
        return [F[R[logical]][periods[fy]] / 1e6 for fy in fys]

    wb = Workbook()

    def header(ws, title, unit_note):
        ws["A1"] = title
        ws["A1"].font = Font(name="Arial", size=12, bold=True)
        ws["A2"] = unit_note
        ws["A2"].font = NOTE
        for i, fy in enumerate(fys):
            c = ws.cell(row=4, column=2 + i, value=fy)
            c.font = BOLD
            c.alignment = RIGHT
        ws.freeze_panes = "B5"
        ws.column_dimensions["A"].width = 34
        for cl in cols:
            ws.column_dimensions[cl].width = 13
        ws.column_dimensions[notes_col].width = 60

    def put(ws, r, label, values=None, formulas=None, font=BLACK, fmt=MONEY,
            note=None, bold=False, border=None):
        ws.cell(row=r, column=1, value=label).font = BOLD if bold else Font(name="Arial", size=11)
        for i in range(n):
            c = ws.cell(row=r, column=2 + i)
            if formulas is not None:
                if formulas[i] is not None:
                    c.value = formulas[i]
            elif values is not None:
                c.value = round(values[i], 1)
            c.font = font
            c.number_format = fmt
            c.alignment = RIGHT
            if border:
                c.border = border
        if note:
            ws.cell(row=r, column=2 + n, value=note).font = NOTE

    gaps = list(meta.get("unresolved_metrics", []))

    # ---------------- Historical Financials ----------------
    ws = wb.active
    ws.title = "Historical Financials"
    header(ws, f"{meta['entity']} -- Income Statement",
           "$ in millions; fiscal years per company reporting calendar")
    r, IS = 5, {}
    if has("revenue"):
        IS["rev"] = r; put(ws, r, "Revenue", values=vals("revenue"), font=BLUE, note=src); r += 1
    if has("revenue") and has("gross_profit"):
        IS["cogs"] = r
        put(ws, r, "Cost of sales",
            formulas=[f"={c}{IS['rev']}-{c}{r+1}" for c in cols],
            note="Derived: Revenue less Gross profit"); r += 1
        IS["gp"] = r
        put(ws, r, "Gross profit", values=vals("gross_profit"), font=BLUE, note=src,
            bold=True, border=TOP); r += 1
        IS["gm"] = r
        put(ws, r, "  Gross margin %",
            formulas=[f"={c}{IS['gp']}/{c}{IS['rev']}" for c in cols], fmt=PCT); r += 2
    if has("rnd"):
        IS["rnd"] = r; put(ws, r, "Research & development", values=vals("rnd"), font=BLUE, note=src); r += 1
    if has("operating_income"):
        if has("gross_profit") and has("rnd"):
            IS["sga"] = r
            put(ws, r, "SG&A and other opex",
                formulas=[f"={c}{IS['gp']}-{c}{IS['rnd']}-{c}{r+1}" for c in cols],
                note="Derived: Gross profit less R&D less Operating income"); r += 1
        IS["opinc"] = r
        put(ws, r, "Operating income", values=vals("operating_income"), font=BLUE, note=src,
            bold=True, border=TOP); r += 1
        if has("revenue"):
            IS["opm"] = r
            put(ws, r, "  Operating margin %",
                formulas=[f"={c}{IS['opinc']}/{c}{IS['rev']}" for c in cols], fmt=PCT); r += 2
    if has("net_income"):
        IS["ni"] = r
        put(ws, r, "Net income", values=vals("net_income"), font=BLUE, note=src,
            bold=True, border=DBL); r += 1
        if has("revenue"):
            IS["nim"] = r
            put(ws, r, "  Net margin %",
                formulas=[f"={c}{IS['ni']}/{c}{IS['rev']}" for c in cols], fmt=PCT); r += 2
    if has("revenue"):
        IS["growth"] = r
        put(ws, r, "  Revenue growth %",
            formulas=[None] + [f"={cols[i]}{IS['rev']}/{cols[i-1]}{IS['rev']}-1" for i in range(1, n)],
            fmt=PCT, note=f"{fys[0]} growth n/a (prior year not presented)"); r += 1

    # ---------------- Balance Sheet (selected) ----------------
    ws = wb.create_sheet("Balance Sheet")
    header(ws, f"{meta['entity']} -- Balance Sheet (selected items)",
           "$ in millions; selected working-capital items only")
    r, BS = 5, {}
    for label, logical in [("Cash and equivalents", "cash"),
                           ("Accounts receivable, net", "ar"),
                           ("Inventory, net", "inventory")]:
        if has(logical):
            BS[logical] = r; put(ws, r, label, values=vals(logical), font=BLUE, note=src); r += 1
        else:
            gaps.append(logical)
    if BS:
        first, last = min(BS.values()), max(BS.values())
        put(ws, r, "Selected items total",
            formulas=[f"=SUM({c}{first}:{c}{last})" for c in cols], bold=True, border=TOP)
        BS["total"] = r

    # ---------------- Cash Flow (selected) ----------------
    ws = wb.create_sheet("Cash Flow")
    header(ws, f"{meta['entity']} -- Cash Flow (selected items)", "$ in millions")
    r, CF = 5, {}
    if has("ocf"):
        CF["ocf"] = r
        put(ws, r, "Net cash from operating activities", values=vals("ocf"), font=BLUE,
            note=src, bold=True); r += 1
        if has("capex"):
            CF["capex"] = r
            put(ws, r, "Less: Capex", values=[-x for x in vals("capex")], font=BLUE, note=src); r += 1
            CF["fcf"] = r
            put(ws, r, "Free cash flow",
                formulas=[f"={c}{CF['ocf']}+{c}{CF['capex']}" for c in cols],
                bold=True, border=TOP); r += 2
        if "ni" in IS:
            CF["ni_link"] = r
            put(ws, r, "  Memo: Net income (link)",
                formulas=[f"='Historical Financials'!{c}{IS['ni']}" for c in cols], font=GREEN); r += 1
            CF["conv"] = r
            put(ws, r, "  OCF / Net income conversion",
                formulas=[f"={c}{CF['ocf']}/{c}{CF['ni_link']}" for c in cols], fmt=PCT); r += 1

    # ---------------- Operating Metrics ----------------
    ws = wb.create_sheet("Operating Metrics")
    header(ws, f"{meta['entity']} -- Operating Metrics", "Ratios computed from statement tabs")
    r = 5
    if "ar" in BS and "rev" in IS:
        put(ws, r, "Days sales outstanding (DSO)",
            formulas=[f"='Balance Sheet'!{c}{BS['ar']}/'Historical Financials'!{c}{IS['rev']}*365" for c in cols],
            font=GREEN, fmt=DAYS, note="Period-end AR / revenue x 365"); r += 1
    if "inventory" in BS and "cogs" in IS:
        put(ws, r, "Days inventory outstanding (DIO)",
            formulas=[f"='Balance Sheet'!{c}{BS['inventory']}/'Historical Financials'!{c}{IS['cogs']}*365" for c in cols],
            font=GREEN, fmt=DAYS, note="Period-end inventory / COGS x 365"); r += 1
    if "rnd" in IS and "rev" in IS:
        put(ws, r, "R&D % of revenue",
            formulas=[f"='Historical Financials'!{c}{IS['rnd']}/'Historical Financials'!{c}{IS['rev']}" for c in cols],
            font=GREEN, fmt=PCT); r += 1

    # ---------------- Executive Summary ----------------
    ws = wb.create_sheet("Executive Summary", 0)
    ws["A1"] = f"{meta['entity']} ({ticker}) -- Data Pack"
    ws["A1"].font = Font(name="Arial", size=13, bold=True)
    ws["A2"] = f"$ in millions unless noted. Source: {src}. Auto-generated; audited by Basis."
    ws["A2"].font = NOTE
    ws.column_dimensions["A"].width = 30
    for cl in cols:
        ws.column_dimensions[cl].width = 13
    ws["A4"] = "Financial snapshot"
    ws["A4"].font = BOLD
    for i, fy in enumerate(fys):
        c = ws.cell(row=5, column=2 + i, value=fy)
        c.font = BOLD; c.alignment = RIGHT
    snap = [("Revenue", "Historical Financials", IS.get("rev"), MONEY),
            ("Revenue growth %", "Historical Financials", IS.get("growth"), PCT),
            ("Gross margin %", "Historical Financials", IS.get("gm"), PCT),
            ("Operating income", "Historical Financials", IS.get("opinc"), MONEY),
            ("Net income", "Historical Financials", IS.get("ni"), MONEY),
            ("Operating cash flow", "Cash Flow", CF.get("ocf"), MONEY),
            ("Free cash flow", "Cash Flow", CF.get("fcf"), MONEY),
            ("Cash and equivalents", "Balance Sheet", BS.get("cash"), MONEY)]
    rr = 6
    for label, sheet, row, fmt in snap:
        if row is None:
            continue
        ws.cell(row=rr, column=1, value=label).font = Font(name="Arial", size=11)
        for i in range(n):
            c = ws.cell(row=rr, column=2 + i, value=f"='{sheet}'!{cols[i]}{row}")
            c.font = GREEN; c.number_format = fmt; c.alignment = RIGHT
        rr += 1

    # Narrative highlights: numbers restated in prose (the audit flags these
    # as NARRATIVE_UNAUDITED by design -- prose bypasses cell-level audit).
    def musd(x):
        return f"${x/1e3:,.1f}B" if abs(x) >= 1e3 else f"${x:,.0f}M"
    lines = []
    if has("revenue"):
        rv = vals("revenue")
        g = rv[-1] / rv[-2] - 1 if len(rv) > 1 and rv[-2] else None
        lines.append(f"{fys[-1]} revenue of {musd(rv[-1])}" + (f" ({g:+.1%} YoY)." if g is not None else "."))
    if has("gross_profit") and has("revenue"):
        lines.append(f"Gross margin of {vals('gross_profit')[-1]/vals('revenue')[-1]:.1%} in {fys[-1]}.")
    if has("ocf"):
        lines.append(f"Operating cash flow of {musd(vals('ocf')[-1])} in {fys[-1]}.")
    if lines:
        ws.cell(row=rr + 1, column=1, value="Highlights: " + " ".join(lines)).font = Font(name="Arial", size=10)
    if gaps:
        ws.cell(row=rr + 3, column=1,
                value=f"Data gaps (no usable XBRL series): {', '.join(sorted(set(gaps)))}").font = NOTE

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import sys
    p = build_datapack(sys.argv[1], sys.argv[2])
    print(f"wrote {p}")
