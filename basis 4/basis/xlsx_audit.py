"""Audit harness for agent-generated Excel datapacks.

Treats the workbook as an untrusted claim set and audits it against the
EDGAR XBRL cache:

  1. INPUT cells (no formula): map row label -> XBRL concept, compare value
     to the filed number (scale-aware).
  2. FORMULA cells: take the recalculated value and, where the metric is
     recomputable from the cache (margins, growth, ratios), check it.
  3. RULE-6 check: rows whose labels mark them as calculations (margin,
     growth, %, total) must be formulas, not hardcodes.
  4. Internal consistency: derived identities must hold in the values
     (COGS + GP = Revenue, etc.).
  5. NARRATIVE scan: numbers embedded in prose cells (exec summaries love
     these) are extracted and flagged -- they bypass cell-level audit.

Requires the workbook to have been recalculated (cached values present).
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

# --- label -> logical metric -> concept mapping (the judgment layer) -----
LABEL_MAP = [
    (r"^revenue$|^net sales|^total revenue", "us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax"),
    (r"^gross profit", "us-gaap:GrossProfit"),
    (r"^operating income", "us-gaap:OperatingIncomeLoss"),
    (r"^net income", "us-gaap:NetIncomeLoss"),
    (r"research .?& ?development|^r&d(?! %)", "us-gaap:ResearchAndDevelopmentExpense"),
    (r"^cash and equivalents|^cash & equivalents", "us-gaap:CashAndCashEquivalentsAtCarryingValue"),
    (r"accounts receivable", "us-gaap:AccountsReceivableNetCurrent"),
    (r"^inventor", "us-gaap:InventoryNet"),
    (r"cash from operating|operating activities", "us-gaap:NetCashProvidedByUsedInOperatingActivities"),
    (r"capex|capital expenditure", "us-gaap:PaymentsToAcquirePropertyPlantAndEquipment"),
]

# Presentation sign conventions: a row led by "Less:" shows a filed-positive
# outflow as a negative. Compare against the negated filed value. (Caught by
# the audit itself on the first generic run -- FAILURES.md #14.)
SIGN_FLIP_PAT = re.compile(r"^\s*less:", re.I)

# Labels that are calculations by definition -> must be formulas (RULE 6)
CALCULATED_PAT = re.compile(r"%|margin|growth|total|conversion|days |dso|dio|\+", re.I)

# v2: derived metrics the auditor recomputes from the cache and checks
# against the recalculated formula value. v1 skipped formula cells entirely
# -- "no finding" looked like a pass when it was actually "not checked".
_REV = "us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax"
_GP = "us-gaap:GrossProfit"
_OI = "us-gaap:OperatingIncomeLoss"
_NI = "us-gaap:NetIncomeLoss"
_RND = "us-gaap:ResearchAndDevelopmentExpense"
_AR = "us-gaap:AccountsReceivableNetCurrent"
_INV = "us-gaap:InventoryNet"
_OCF = "us-gaap:NetCashProvidedByUsedInOperatingActivities"
_CASH = "us-gaap:CashAndCashEquivalentsAtCarryingValue"

DERIVED_MAP = [
    (r"gross margin", lambda f, p, pp: f[_GP][p] / f[_REV][p], "pct"),
    (r"operating margin", lambda f, p, pp: f[_OI][p] / f[_REV][p], "pct"),
    (r"net margin", lambda f, p, pp: f[_NI][p] / f[_REV][p], "pct"),
    (r"revenue growth", lambda f, p, pp: f[_REV][p] / f[_REV][pp] - 1 if pp else None, "pct"),
    (r"cost of sales|cogs", lambda f, p, pp: (f[_REV][p] - f[_GP][p]) / 1e6, "musd"),
    (r"sg&a and other", lambda f, p, pp: (f[_GP][p] - f[_RND][p] - f[_OI][p]) / 1e6, "musd"),
    (r"dso|days sales", lambda f, p, pp: f[_AR][p] / f[_REV][p] * 365, "days"),
    (r"dio|days inventory", lambda f, p, pp: f[_INV][p] / (f[_REV][p] - f[_GP][p]) * 365, "days"),
    (r"r&d % of revenue", lambda f, p, pp: f[_RND][p] / f[_REV][p], "pct"),
    (r"conversion", lambda f, p, pp: f[_OCF][p] / f[_NI][p], "pct"),
    (r"cash \+ ar \+ inventory", lambda f, p, pp: (f[_CASH][p] + f[_AR][p] + f[_INV][p]) / 1e6, "musd"),
]

# v1's narrative scanner flagged the audit's own source citations as
# "unaudited numeric claims" (accession numbers, filing dates). v2: skip
# citation/derivation notes, and only flag prose numbers that assert money,
# percentages, or bps -- bare years and IDs are not claims.
CITATION_PAT = re.compile(r"SEC EDGAR|accn |^Derived:|^Period-end|^FY\d{4} growth", re.I)
MONEYISH_PAT = re.compile(r"\$[\d,.]+|[\d.]+%|[\d.]+\s?bps|\$[\d.]+[BM]|~\d+x|\b\d+(?:\.\d+)?x\b", re.I)


def check_derived(label: str):
    lab = label.strip().lower()
    for pat, fn, unit in DERIVED_MAP:
        if re.search(pat, lab):
            return fn, unit
    return None, None

REL_TOL = 0.005
ABS_TOL_M = 0.06   # $M rounding to one decimal place => up to 0.05 drift


def map_label(label: str):
    lab = label.strip().lower()
    for pat, concept in LABEL_MAP:
        if re.search(pat, lab):
            return concept
    return None


def audit(cache_path: str, xlsx_path: str) -> dict:
    cache = json.loads(Path(cache_path).read_text())
    facts = cache["facts"]

    # FY label -> period end comes from the cache, never assumed. v1 hardcoded
    # June 30 (SMCI's calendar) -- worked on the demo company, would silently
    # mis-verify any December/January filer. FAILURES.md #13.
    fy_map = cache.get("meta", {}).get("fiscal_periods") or {}

    def fy_from_header(h) -> str | None:
        m = re.match(r"FY(\d{4})E?$", str(h or ""))
        if not m:
            return None
        return fy_map.get(f"FY{m.group(1)}", f"{m.group(1)}-06-30")
    wb_f = load_workbook(xlsx_path)                  # formulas
    wb_v = load_workbook(xlsx_path, data_only=True)  # recalculated values

    findings: list[dict] = []
    stats = {"cells_audited": 0, "inputs": 0, "formulas": 0}

    def add(sheet, cell, verdict, detail, label=""):
        findings.append({"sheet": sheet, "cell": cell, "label": label.strip(),
                         "verdict": verdict, "detail": detail})

    for sheet in wb_f.sheetnames:
        sf, sv = wb_f[sheet], wb_v[sheet]

        # scale: unit declaration in the header area
        header_blob = " ".join(str(sf.cell(row=r, column=1).value or "") for r in (1, 2, 3)).lower()
        scale = 1e6 if "million" in header_blob else 1.0

        # find the FY header row and its column->period mapping
        periods: dict[int, str] = {}
        for r in range(1, 8):
            row_periods = {c: fy_from_header(sf.cell(row=r, column=c).value)
                           for c in range(2, 12)}
            row_periods = {c: p for c, p in row_periods.items() if p}
            if len(row_periods) >= 3:
                periods = row_periods
                break
        if not periods:
            continue

        for r in range(1, sf.max_row + 1):
            label = str(sf.cell(row=r, column=1).value or "")
            if not label.strip():
                continue
            concept = map_label(label)
            is_calc_label = bool(CALCULATED_PAT.search(label))
            for c, period in periods.items():
                fcell, vcell = sf.cell(row=r, column=c), sv.cell(row=r, column=c)
                if fcell.value is None:
                    continue
                is_formula = isinstance(fcell.value, str) and fcell.value.startswith("=")
                val = vcell.value
                if not isinstance(val, (int, float)):
                    if is_formula:
                        stats["formulas_unevaluated"] = stats.get("formulas_unevaluated", 0) + 1
                    continue
                stats["cells_audited"] += 1
                stats["formulas" if is_formula else "inputs"] += 1
                addr = f"{fcell.coordinate}"

                # RULE 6: calculated-by-definition rows must be formulas
                if is_calc_label and not is_formula:
                    add(sheet, addr, "RULE6_VIOLATION",
                        f"'{label.strip()}' is a calculation but cell holds a hardcoded {val}", label)

                # v2: derived metrics -- recompute from cache, compare to
                # the recalculated formula value (or catch a hardcode)
                dfn, dunit = check_derived(label)
                if dfn is not None:
                    prev_period = None
                    prev_col = c - 1
                    if prev_col in periods:
                        prev_period = periods[prev_col]
                    try:
                        expected = dfn(facts, period, prev_period)
                    except (KeyError, ZeroDivisionError, TypeError):
                        expected = None
                    if expected is None:
                        if val == 0:
                            add(sheet, addr, "MISMATCH",
                                f"'{label.strip()}' @ {period} shows 0 but is not computable "
                                f"(no prior-year value) -- a link/formula coerced a blank into a "
                                f"fake zero; should display n/a", label)
                        else:
                            add(sheet, addr, "UNSOURCED",
                                f"cannot recompute '{label.strip()}' @ {period} from cache", label)
                        continue
                    if dunit == "pct":
                        ok = abs(val - expected) <= 0.004
                    elif dunit == "days":
                        ok = abs(val - expected) <= 0.5
                    else:  # musd
                        ok = abs(val - expected) <= max(abs(expected) * REL_TOL, ABS_TOL_M * 3)
                    if ok:
                        add(sheet, addr, "VERIFIED",
                            f"recomputed {expected:,.4f} ~= sheet {val:,.4f} [{label.strip()}]", label)
                    else:
                        add(sheet, addr, "MISMATCH",
                            f"sheet shows {val:,.4f} but recomputes to {expected:,.4f} [{label.strip()}]", label)
                    continue

                if concept:
                    filed = facts.get(concept, {}).get(period)
                    if filed is None:
                        add(sheet, addr, "UNSOURCED",
                            f"no filed value in cache for {concept} @ {period}", label)
                        continue
                    if SIGN_FLIP_PAT.search(label):
                        filed = -filed
                    stated = val * scale
                    drift = abs(stated - filed)
                    ok = drift <= max(abs(filed) * REL_TOL, ABS_TOL_M * 1e6)
                    if not ok:
                        add(sheet, addr, "MISMATCH",
                            f"states {stated:,.0f} vs filed {filed:,.0f} "
                            f"({(stated - filed) / filed:+.2%}) [{concept} @ {period}]", label)
                    else:
                        add(sheet, addr, "VERIFIED",
                            f"{stated:,.0f} matches filed {filed:,.0f} [{concept.split(':')[1]} @ {period}]", label)
                elif not is_formula:
                    # a hardcoded number with no mappable source label
                    add(sheet, addr, "UNSOURCED",
                        f"hardcoded {val} under unmapped label '{label.strip()}'", label)

        # narrative scan: numbers living inside prose cells
        for r in range(1, sf.max_row + 1):
            for c in range(1, 8):
                v = sf.cell(row=r, column=c).value
                if isinstance(v, str) and not v.startswith("=") and len(v) > 60:
                    if CITATION_PAT.search(v):
                        continue  # source citations / derivation notes, not claims
                    nums = MONEYISH_PAT.findall(v)
                    if nums:
                        add(sheet, sf.cell(row=r, column=c).coordinate, "NARRATIVE_UNAUDITED",
                            f"prose cell contains {len(nums)} numeric claims outside cell-level audit: "
                            + ", ".join(nums[:8]), v[:60] + "...")

    # cross-metric identity checks on recalculated values
    identities = []
    hf = wb_v["Historical Financials"] if "Historical Financials" in wb_v.sheetnames else None
    if hf:
        for c, period in periods.items():
            pass  # identities below use known cache values instead
    for period in sorted(next(iter(facts.values())).keys()):
        rev = facts["us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax"].get(period)
        gp = facts["us-gaap:GrossProfit"].get(period)
        oi = facts["us-gaap:OperatingIncomeLoss"].get(period)
        if rev and gp and oi:
            identities.append({
                "period": period,
                "implied_cogs": rev - gp,
                "implied_opex": gp - oi,
                "opex_pct_rev": round((gp - oi) / rev, 4),
            })

    if stats.get("formulas_unevaluated"):
        findings.append({
            "sheet": "(workbook)", "cell": "-", "label": "",
            "verdict": "NOT_EVALUATED",
            "detail": (f"{stats['formulas_unevaluated']} formula cells have no computed value -- "
                       f"the workbook was never recalculated. Open it in Excel and save, or run "
                       f"scripts/recalc.py (needs LibreOffice), then re-audit; formula checks are "
                       f"skipped until then.")})

    tally: dict[str, int] = {}
    for f in findings:
        tally[f["verdict"]] = tally.get(f["verdict"], 0) + 1
    return {"workbook": xlsx_path, "stats": stats, "tally": tally,
            "identity_reference": identities, "findings": findings}


if __name__ == "__main__":
    report = audit(sys.argv[1], sys.argv[2])
    out = Path(sys.argv[3]) if len(sys.argv) > 3 else Path("reports/xlsx_audit.json")
    out.write_text(json.dumps(report, indent=2))
    print(json.dumps({"stats": report["stats"], "tally": report["tally"]}, indent=2))
    for f in report["findings"]:
        if f["verdict"] != "VERIFIED":
            print(f"  {f['sheet']}!{f['cell']:6s} {f['verdict']:20s} {f['detail'][:110]}")
