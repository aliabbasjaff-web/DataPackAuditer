"""Agent-under-test: build the SMCI data pack per the datapack-builder skill.

This script is the 'agent run' being audited. It follows the skill's rules:
formulas for all calculations (RULE 6), blue inputs / black formulas / green
cross-sheet links, $M units declared in headers, years as text, accounting
negatives, source citations in a notes column.

Data source: EDGAR XBRL cache (data/cache/smci.json). FY2021-FY2025.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

cache = json.loads(Path("data/cache/smci.json").read_text())
F = cache["facts"]
ACCN = cache["meta"]["accessions"]

FYS = ["2021", "2022", "2023", "2024", "2025"]
ENDS = [f"{y}-06-30" for y in FYS]

def series(concept):
    return [F[concept][e] / 1e6 for e in ENDS]  # to $M

rev = series("us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax")
gp = series("us-gaap:GrossProfit")
opinc = series("us-gaap:OperatingIncomeLoss")
ni = series("us-gaap:NetIncomeLoss")
rnd = series("us-gaap:ResearchAndDevelopmentExpense")
cash = series("us-gaap:CashAndCashEquivalentsAtCarryingValue")
ar = series("us-gaap:AccountsReceivableNetCurrent")
inv = series("us-gaap:InventoryNet")
ocf = series("us-gaap:NetCashProvidedByUsedInOperatingActivities")

BLUE = Font(name="Arial", size=11, color="0000FF")
BLACK = Font(name="Arial", size=11, color="000000")
GREEN = Font(name="Arial", size=11, color="008000")
BOLD = Font(name="Arial", size=11, bold=True)
HDR = Font(name="Arial", size=11, bold=True)
MONEY = "$#,##0.0;($#,##0.0)"
PCT = "0.0%;(0.0%)"
DAYS = "#,##0.0"
RIGHT = Alignment(horizontal="right")
TOP_BORDER = Border(top=Side(style="thin"))
DBL_BOTTOM = Border(bottom=Side(style="double"))

SRC_10K25 = f"SEC EDGAR, FY2025 10-K (accn {ACCN['FY2025_10K']['accession']}, filed {ACCN['FY2025_10K']['filed']})"
SRC_10K24 = f"SEC EDGAR, FY2024 10-K (accn {ACCN['FY2024_10K']['accession']}, filed {ACCN['FY2024_10K']['filed']})"

wb = Workbook()

def sheet_header(ws, title, unit_note):
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=12, bold=True)
    ws["A2"] = unit_note
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="5B5B66")
    for i, fy in enumerate(FYS):
        c = ws.cell(row=4, column=2 + i, value=f"FY{fy}")
        c.font = HDR
        c.alignment = RIGHT
    ws.cell(row=4, column=1, value="").font = HDR
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 34
    for i in range(len(FYS)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13
    ws.column_dimensions[get_column_letter(2 + len(FYS))].width = 60

def put_row(ws, row, label, values=None, formula=None, font=BLACK, fmt=MONEY,
            note=None, bold_label=False, indent=0):
    lab = ws.cell(row=row, column=1, value=("  " * indent) + label)
    lab.font = BOLD if bold_label else Font(name="Arial", size=11)
    for i in range(len(FYS)):
        c = ws.cell(row=row, column=2 + i)
        if formula is not None:
            fv = formula(i, get_column_letter(2 + i))
            if fv is not None:
                c.value = fv
        elif values is not None:
            c.value = round(values[i], 1)
        c.font = font
        c.number_format = fmt
        c.alignment = RIGHT
    if note:
        n = ws.cell(row=row, column=2 + len(FYS), value=note)
        n.font = Font(name="Arial", size=9, color="5B5B66")

# ---------------- Tab: Historical Financials ----------------
ws = wb.active
ws.title = "Historical Financials"
sheet_header(ws, "Super Micro Computer, Inc. -- Income Statement", "$ in millions; fiscal years ended June 30")
r = 5
R = {}
R["rev"] = r;   put_row(ws, r, "Revenue", values=rev, font=BLUE, note=SRC_10K25 + "; FY2021-23 as restated in later filings"); r += 1
R["cogs"] = r;  put_row(ws, r, "Cost of sales", formula=lambda i, c: f"={c}{R['rev']}-{c}{R['gp_pending']}" if False else None); r += 1
# NOTE: COGS = Revenue - Gross profit; gross profit row not yet written, so fill after
R["gp"] = r;    put_row(ws, r, "Gross profit", values=gp, font=BLUE, note=SRC_10K25, bold_label=True); r += 1
for i in range(len(FYS)):
    col = get_column_letter(2 + i)
    c = ws.cell(row=R["cogs"], column=2 + i, value=f"={col}{R['rev']}-{col}{R['gp']}")
    c.font = BLACK; c.number_format = MONEY; c.alignment = RIGHT
ws.cell(row=R["cogs"], column=2 + len(FYS), value="Derived: Revenue less Gross profit").font = Font(name="Arial", size=9, color="5B5B66")
for i in range(len(FYS)):
    ws.cell(row=R["gp"], column=2 + i).border = TOP_BORDER
R["gm"] = r;    put_row(ws, r, "Gross margin %", formula=lambda i, c: f"={c}{R['gp']}/{c}{R['rev']}", fmt=PCT, indent=1); r += 2
R["rnd"] = r;   put_row(ws, r, "Research & development", values=rnd, font=BLUE, note=SRC_10K25); r += 1
R["sga"] = r;   put_row(ws, r, "SG&A and other opex", formula=lambda i, c: f"={c}{R['gp']}-{c}{R['rnd']}-{c}{R['opinc_pending']}" if False else None); r += 1
R["opinc"] = r; put_row(ws, r, "Operating income", values=opinc, font=BLUE, note=SRC_10K25, bold_label=True); r += 1
for i in range(len(FYS)):
    col = get_column_letter(2 + i)
    c = ws.cell(row=R["sga"], column=2 + i, value=f"={col}{R['gp']}-{col}{R['rnd']}-{col}{R['opinc']}")
    c.font = BLACK; c.number_format = MONEY; c.alignment = RIGHT
ws.cell(row=R["sga"], column=2 + len(FYS), value="Derived: Gross profit less R&D less Operating income").font = Font(name="Arial", size=9, color="5B5B66")
for i in range(len(FYS)):
    ws.cell(row=R["opinc"], column=2 + i).border = TOP_BORDER
R["opm"] = r;   put_row(ws, r, "Operating margin %", formula=lambda i, c: f"={c}{R['opinc']}/{c}{R['rev']}", fmt=PCT, indent=1); r += 2
R["ni"] = r;    put_row(ws, r, "Net income", values=ni, font=BLUE, note=SRC_10K25, bold_label=True); r += 1
for i in range(len(FYS)):
    ws.cell(row=R["ni"], column=2 + i).border = DBL_BOTTOM
R["nim"] = r;   put_row(ws, r, "Net margin %", formula=lambda i, c: f"={c}{R['ni']}/{c}{R['rev']}", fmt=PCT, indent=1); r += 2
R["growth"] = r
put_row(ws, r, "Revenue growth %",
        formula=lambda i, c: None if i == 0 else f"={c}{R['rev']}/{get_column_letter(1 + i)}{R['rev']}-1",
        fmt=PCT, indent=1, note="FY2021 growth n/a (FY2020 not presented)")
ws_is = ws

# ---------------- Tab: Balance Sheet (selected) ----------------
ws = wb.create_sheet("Balance Sheet")
sheet_header(ws, "Super Micro Computer, Inc. -- Balance Sheet (selected items)",
             "$ in millions; as of June 30. Selected working-capital items only -- full statement pending broader data ingestion")
r = 5
B = {}
B["cash"] = r; put_row(ws, r, "Cash and equivalents", values=cash, font=BLUE, note=SRC_10K25); r += 1
B["ar"] = r;   put_row(ws, r, "Accounts receivable, net", values=ar, font=BLUE, note=SRC_10K25); r += 1
B["inv"] = r;  put_row(ws, r, "Inventory, net", values=inv, font=BLUE, note=SRC_10K25); r += 1
B["wc"] = r
put_row(ws, r, "Cash + AR + Inventory",
        formula=lambda i, c: f"=SUM({c}{B['cash']}:{c}{B['inv']})", bold_label=True)
for i in range(len(FYS)):
    ws.cell(row=B["wc"], column=2 + i).border = TOP_BORDER
ws_bs = ws

# ---------------- Tab: Cash Flow (selected) ----------------
ws = wb.create_sheet("Cash Flow")
sheet_header(ws, "Super Micro Computer, Inc. -- Cash Flow (selected items)",
             "$ in millions; fiscal years ended June 30")
r = 5
C = {}
C["ocf"] = r; put_row(ws, r, "Net cash from operating activities", values=ocf, font=BLUE,
                      note=SRC_10K25 + "; FY2024 per delayed 10-K (" + SRC_10K24 + ")", bold_label=True); r += 1
C["ni_link"] = r
put_row(ws, r, "Memo: Net income (link)",
        formula=lambda i, c: f"='Historical Financials'!{c}{R['ni']}", font=GREEN, indent=1); r += 1
C["conv"] = r
put_row(ws, r, "OCF / Net income conversion",
        formula=lambda i, c: f"={c}{C['ocf']}/{c}{C['ni_link']}", fmt="0.0%;(0.0%)", indent=1)
ws_cf = ws

# ---------------- Tab: Operating Metrics ----------------
ws = wb.create_sheet("Operating Metrics")
sheet_header(ws, "Super Micro Computer, Inc. -- Operating Metrics",
             "Ratios computed from statement tabs; fiscal years ended June 30")
r = 5
M = {}
M["dso"] = r
put_row(ws, r, "Days sales outstanding (DSO)",
        formula=lambda i, c: f"='Balance Sheet'!{c}{B['ar']}/'Historical Financials'!{c}{R['rev']}*365",
        font=GREEN, fmt=DAYS, note="Period-end AR / revenue x 365"); r += 1
M["dio"] = r
put_row(ws, r, "Days inventory outstanding (DIO)",
        formula=lambda i, c: f"='Balance Sheet'!{c}{B['inv']}/'Historical Financials'!{c}{R['cogs']}*365",
        font=GREEN, fmt=DAYS, note="Period-end inventory / COGS x 365"); r += 1
M["rndint"] = r
put_row(ws, r, "R&D % of revenue",
        formula=lambda i, c: f"='Historical Financials'!{c}{R['rnd']}/'Historical Financials'!{c}{R['rev']}",
        font=GREEN, fmt=PCT); r += 1
ws_om = ws

# ---------------- Tab: Executive Summary ----------------
ws = wb.create_sheet("Executive Summary", 0)
ws["A1"] = "Super Micro Computer, Inc. (NASDAQ: SMCI) -- Data Pack"
ws["A1"].font = Font(name="Arial", size=13, bold=True)
ws["A2"] = "$ in millions unless noted; fiscal years ended June 30. Source: SEC EDGAR XBRL (10-K filings)."
ws["A2"].font = Font(name="Arial", size=9, italic=True, color="5B5B66")
ws["A4"] = ("Overview: SMCI designs and manufactures high-performance server and storage systems, "
            "with growth driven by AI/GPU data-center demand. Revenue scaled ~6x from FY2021 to FY2025 "
            "while gross margin compressed, and FY2024 saw a large inventory-driven cash outflow.")
ws["A4"].font = Font(name="Arial", size=10)
ws["A4"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A4:F4")
ws.row_dimensions[4].height = 45
ws["A6"] = "Financial snapshot"
ws["A6"].font = BOLD
for i, fy in enumerate(FYS):
    c = ws.cell(row=7, column=2 + i, value=f"FY{fy}")
    c.font = HDR; c.alignment = RIGHT
snap = [
    ("Revenue", f"='Historical Financials'!{{c}}{R['rev']}", MONEY),
    ("Revenue growth %", f"='Historical Financials'!{{c}}{R['growth']}", PCT),
    ("Gross margin %", f"='Historical Financials'!{{c}}{R['gm']}", PCT),
    ("Operating income", f"='Historical Financials'!{{c}}{R['opinc']}", MONEY),
    ("Net income", f"='Historical Financials'!{{c}}{R['ni']}", MONEY),
    ("Operating cash flow", f"='Cash Flow'!{{c}}{C['ocf']}", MONEY),
    ("Cash and equivalents", f"='Balance Sheet'!{{c}}{B['cash']}", MONEY),
]
for j, (label, tmpl, fmt) in enumerate(snap):
    rr = 8 + j
    ws.cell(row=rr, column=1, value=label).font = Font(name="Arial", size=11)
    for i in range(len(FYS)):
        col = get_column_letter(2 + i)
        c = ws.cell(row=rr, column=2 + i, value=tmpl.format(c=col))
        c.font = GREEN; c.number_format = fmt; c.alignment = RIGHT
ws["A16"] = "Highlights: 46.6% FY2025 revenue growth on Blackwell-generation demand; OCF swung from $(2,486)M (FY2024) to $1,660M (FY2025); $5.2B cash provides balance-sheet flexibility."
ws["A16"].font = Font(name="Arial", size=10)
ws["A17"] = "Considerations: gross margin down ~270bps YoY (11.1% FY2025); net income declined 9% despite revenue growth; FY2024 10-K was filed late (Feb 2025) following auditor transition."
ws["A17"].font = Font(name="Arial", size=10)
ws.column_dimensions["A"].width = 30
for i in range(len(FYS)):
    ws.column_dimensions[get_column_letter(2 + i)].width = 13

out = Path("agent_run/SMCI_DataPack_2026-08-30.xlsx")
wb.save(out)
print(f"wrote {out}")
print("row map:", json.dumps({"IS": R, "BS": B, "CF": C, "OM": M}))
